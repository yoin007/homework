from datetime import date, datetime, timedelta
from typing import List, Optional, Dict, Any
import logging

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from starlette.background import BackgroundTask
from fastapi.responses import FileResponse
from sqlalchemy import func, and_, or_, select
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
import pandas as pd
import os
import tempfile

from app.core.database import get_db
from app.models.models import Record, Student, Teacher
from app.schemas.schemas import Record as RecordSchema
from app.schemas.schemas import RecordCreate, RecordFilter, StudentScoreSummary, SubjectScoreSummary, ExcelImportResponse, ExcelExportRequest
from app.utils.grade_calculator import GradeCalculator
from openpyxl.styles import PatternFill

router = APIRouter()
logger = logging.getLogger(__name__)


@router.get("/", response_model=List[RecordSchema])
def read_records(
    grade: Optional[str] = None,
    class_name: Optional[str] = None,
    group: Optional[str] = None,
    subject: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """获取作业记录列表，支持筛选"""
    try:
        # 如果任何参数为None，返回空列表
        if any(param is None for param in [grade, start_date, end_date]):
            return []

        filter_params = RecordFilter(
            grade=grade,
            class_name=class_name,
            group=group,
            subject=subject,
            start_date=start_date,
            end_date=end_date
        )

        stmt = select(Record)
        if any([filter_params.grade, filter_params.class_name, filter_params.group]):
            stmt = stmt.join(Student)
            if filter_params.grade:
                stmt = stmt.where(Student.grade == filter_params.grade)
            if filter_params.class_name:
                stmt = stmt.where(Student.class_name == filter_params.class_name)
            if filter_params.group:
                stmt = stmt.where(Student.group == filter_params.group)
        
        if filter_params.subject:
            stmt = stmt.where(Record.subject == filter_params.subject)
        if filter_params.start_date:
            stmt = stmt.where(Record.date >= filter_params.start_date)
        if filter_params.end_date:
            stmt = stmt.where(Record.date <= filter_params.end_date)
        
        records = db.execute(stmt).scalars().all()
        return records
    except SQLAlchemyError as e:
        logger.error(f"查询记录失败: {str(e)}")
        raise HTTPException(status_code=500, detail="数据库查询失败")
    except ValueError as e:
        logger.error(f"参数验证失败: {str(e)}")
        raise HTTPException(status_code=422, detail=str(e))

@router.post("/records", response_model=RecordSchema)
def create_record(record: RecordCreate, db: Session = Depends(get_db)):
    """创建新的作业记录"""
    try:
        # 检查学生是否存在
        student = db.query(Student).filter(Student.student_id == record.student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail=f"学生ID {record.student_id} 不存在")
        
        # 检查教师是否存在
        teacher = db.query(Teacher).filter(Teacher.teacher_id == record.teacher_id).first()
        if not teacher:
            raise HTTPException(status_code=404, detail=f"教师ID {record.teacher_id} 不存在")
        
        db_record = Record(**record.dict())
        db.add(db_record)
        try:
            db.commit()
            db.refresh(db_record)
            logger.info(f"成功创建记录: ID={db_record.id}")
            return db_record
        except SQLAlchemyError as e:
            db.rollback()
            logger.error(f"创建记录失败: {str(e)}")
            raise HTTPException(status_code=500, detail="创建记录失败")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"创建记录时发生错误: {str(e)}")
        raise HTTPException(status_code=500, detail="服务器内部错误")


@router.get("/batches/all", response_model=List[str])
def get_all_batches(db: Session = Depends(get_db)):
    """获取所有批次"""
    try:
        batches = db.query(Record.batch).distinct().all()
        return [batch[0] for batch in batches]
    except SQLAlchemyError as e:
        logger.error(f"Error getting all batches: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@router.get("/batches", response_model=List[str])
def get_batches(db: Session = Depends(get_db)):
    """获取批次列表"""
    try:
        batches = db.query(Record.batch).distinct().all()
        return [batch[0] for batch in batches]
    except SQLAlchemyError as e:
        logger.error(f"Error getting batches: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@router.get("/{record_id}", response_model=RecordSchema)
def read_record(record_id: int, db: Session = Depends(get_db)):
    """
    获取指定ID的作业记录
    """
    record = db.query(Record).filter(Record.id == record_id).first()
    if record is None:
        raise HTTPException(status_code=404, detail=f"作业记录ID {record_id} 不存在")
    return record


@router.put("/{record_id}", response_model=RecordSchema)
def update_record(record_id: int, record: RecordCreate, db: Session = Depends(get_db)):
    """更新指定ID的作业记录"""
    try:
        db_record = db.query(Record).filter(Record.id == record_id).with_for_update().first()
        if db_record is None:
            raise HTTPException(status_code=404, detail=f"作业记录ID {record_id} 不存在")
        
        # 检查学生是否存在
        student = db.query(Student).filter(Student.student_id == record.student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail=f"学生ID {record.student_id} 不存在")
        
        # 检查教师是否存在
        teacher = db.query(Teacher).filter(Teacher.teacher_id == record.teacher_id).first()
        if not teacher:
            raise HTTPException(status_code=404, detail=f"教师ID {record.teacher_id} 不存在")
        
        try:
            # 更新记录
            for key, value in record.dict().items():
                setattr(db_record, key, value)
            
            db.commit()
            db.refresh(db_record)
            logger.info(f"成功更新记录: ID={record_id}")
            return db_record
        except SQLAlchemyError as e:
            db.rollback()
            logger.error(f"更新记录失败: {str(e)}")
            raise HTTPException(status_code=500, detail="更新记录失败")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新记录时发生错误: {str(e)}")
        raise HTTPException(status_code=500, detail="服务器内部错误")

@router.delete("/{record_id}")
def delete_record(record_id: int, db: Session = Depends(get_db)):
    """删除指定ID的作业记录"""
    try:
        db_record = db.query(Record).filter(Record.id == record_id).with_for_update().first()
        if db_record is None:
            raise HTTPException(status_code=404, detail=f"作业记录ID {record_id} 不存在")
        
        try:
            db.delete(db_record)
            db.commit()
            logger.info(f"成功删除记录: ID={record_id}")
            return {"message": f"作业记录ID {record_id} 已删除"}
        except SQLAlchemyError as e:
            db.rollback()
            logger.error(f"删除记录失败: {str(e)}")
            raise HTTPException(status_code=500, detail="删除记录失败")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除记录时发生错误: {str(e)}")
        raise HTTPException(status_code=500, detail="服务器内部错误")


@router.post("/import", response_model=ExcelImportResponse)
def import_records_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """从Excel导入作业记录"""
    # 检查文件名是否以.xls或.xlsx结尾
    if not (file.filename and (file.filename.endswith('.xls') or file.filename.endswith('.xlsx'))):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xls或.xlsx)")
    
    temp_file_path = None
    try:
        # 创建临时文件保存上传的Excel
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_file.write(file.file.read())
            temp_file_path = temp_file.name
        
        # 读取Excel文件
        df = pd.read_excel(temp_file_path)
        
        # 检查必要的列是否存在
        required_columns = ['学号', '姓名', '学科', '日期', '教师ID']
        for col in required_columns:
            if col not in df.columns:
                raise HTTPException(status_code=400, detail=f"Excel文件缺少必要的列: {col}")
        
        # 转换日期列
        df['日期'] = pd.to_datetime(df['日期']).dt.date
        
        # 导入记录
        imported_count = 0
        failed_records = []
        
        try:
            for _, row in df.iterrows():
                try:
                    # 检查学生是否存在
                    student = db.query(Student).filter(Student.student_id == row['学号']).first()
                    if not student:
                        failed_records.append({"学号": row['学号'], "原因": "学生不存在"})
                        continue
                    
                    # 检查教师是否存在
                    teacher = db.query(Teacher).filter(Teacher.teacher_id == row['教师ID']).first()
                    if not teacher:
                        failed_records.append({"学号": row['学号'], "原因": "教师不存在"})
                        continue
                    
                    # 创建记录
                    record_data = {
                        'student_id': row['学号'],
                        'name': row['姓名'],
                        'subject': row['学科'],
                        'date': row['日期'],
                        'teacher_id': row['教师ID'],
                        'score': row.get('分数'),
                        'type': row.get('类型'),
                        'batch': row.get('批次')
                    }
                    
                    db_record = Record(**record_data)
                    db.add(db_record)
                    imported_count += 1
                    
                except Exception as e:
                    failed_records.append({"学号": row['学号'], "原因": str(e)})
                    logger.error(f"导入记录失败: {str(e)}")
                    continue
            
            db.commit()
            logger.info(f"成功导入 {imported_count} 条记录")
            
            return {
                "success": True,
                "message": f"成功导入 {imported_count} 条记录，失败 {len(failed_records)} 条",
                "imported_count": imported_count,
                "failed_records": failed_records
            }
            
        except SQLAlchemyError as e:
            db.rollback()
            logger.error(f"批量导入记录失败: {str(e)}")
            raise HTTPException(status_code=500, detail="批量导入记录失败")
            
    except Exception as e:
        logger.error(f"处理Excel文件失败: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
        
    finally:
        # 确保临时文件被删除
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except Exception as e:
                logger.error(f"删除临时文件失败: {str(e)}")


@router.post("/export-template")
def export_student_template(request: ExcelExportRequest, db: Session = Depends(get_db)):
    """
    导出学生信息模板，用于作业数据录入
    """
    query = db.query(Student)
    
    # 应用筛选条件
    if request.grade:
        query = query.filter(Student.grade == request.grade)
    
    if request.class_name and len(request.class_name) > 0:
        query = query.filter(Student.class_name.in_(request.class_name))
    
    students = query.all()
    
    if not students:
        raise HTTPException(status_code=404, detail="未找到符合条件的学生")
    
    # 创建DataFrame
    data = []
    for student in students:
        data.append({
            "学号": student.student_id,
            "姓名": student.name,
            "班级": student.class_name,
            "学科": "",
            "分数": None,
            "日期": datetime.now().date(),
            "批次": ""
        })
    
    df = pd.DataFrame(data)
    
    # 创建临时文件保存Excel
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        df.to_excel(temp_file.name, index=False)
        temp_file_path = temp_file.name
    
    # 创建后台任务用于删除临时文件
    background_task = BackgroundTask(os.unlink, temp_file_path)
    
    # 返回文件，并设置后台任务
    return FileResponse(
        temp_file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"学生作业模板_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
        background=background_task
    )


@router.post("/summary")
def get_student_score_summary(filter: RecordFilter, db: Session = Depends(get_db)):
    """获取学生成绩汇总"""
    try:
        # 构建查询
        stmt = select(
            Student.student_id,
            Student.name,
            Student.grade,
            Student.class_name,
            Student.group,
            Record.subject,
            func.avg(Record.score).label("avg_score")
        ).join(Student)

        # 应用筛选条件
        if filter.grade:
            stmt = stmt.where(Student.grade == filter.grade)
        if filter.class_name:
            stmt = stmt.where(Student.class_name == filter.class_name)
        if filter.group:
            stmt = stmt.where(Student.group == filter.group)
        if filter.subject:
            stmt = stmt.where(Record.subject == filter.subject)
        if filter.start_date:
            stmt = stmt.where(Record.date >= filter.start_date)
        if filter.end_date:
            stmt = stmt.where(Record.date <= filter.end_date)

        stmt = stmt.group_by(
            Student.student_id,
            Student.name,
            Student.grade,
            Student.class_name,
            Student.group,
            Record.subject
        )

        results = db.execute(stmt).all()

        # 处理结果
        student_scores = {}
        for r in results:
            if r.student_id not in student_scores:
                student_scores[r.student_id] = {
                    "student_id": r.student_id,
                    "name": r.name,
                    "grade": r.grade,
                    "class_name": r.class_name,
                    "group": r.group,
                    "total_score": 0
                }
            student_scores[r.student_id][r.subject] = r.avg_score
            student_scores[r.student_id]["total_score"] += r.avg_score or 0

        # 计算等级并添加颜色
        summary = []
        for i, (_, score) in enumerate(student_scores.items(), 1):
            score["id"] = i
            grade = GradeCalculator.calculate_grade(score["total_score"])
            score["grade"] = grade
            score["grade_color"] = GradeCalculator.get_grade_color(grade)
            summary.append(score)

        return summary
    except SQLAlchemyError as e:
        logger.error(f"查询记录失败: {str(e)}")
        raise HTTPException(status_code=500, detail="数据库查询失败")


@router.post("/export-summary")
def export_student_score_summary_to_excel_v2(filter: RecordFilter, db: Session = Depends(get_db)):
    """导出学生成绩汇总到Excel"""
    try:
        # 获取汇总数据
        summary = get_student_score_summary(filter, db)
        
        # 创建DataFrame
        df = pd.DataFrame(summary)
        
        # 重新排列列顺序
        columns = [
            'student_id', 'name', 'grade', 'class_name', 'group',
            '语文', '数学', '英语',  # 固定学科列
            'total_score', 'grade'  # 总分和等级
        ]
        
        # 确保所有必要的列都存在
        for col in columns:
            if col not in df.columns and col not in ['语文', '数学', '英语']:
                df[col] = None
        
        # 重命名列
        column_names = {
            'student_id': '学号',
            'name': '姓名',
            'grade': '年级',
            'class_name': '班级',
            'group': '小组',
            'total_score': '总分'
        }
        df = df.rename(columns=column_names)
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            # 使用ExcelWriter以便设置单元格样式
            with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='成绩汇总')
                
                # 获取工作表
                worksheet = writer.sheets['成绩汇总']
                
                # 设置等级列的背景颜色
                for idx, row in df.iterrows():
                    grade = row.get('grade')
                    if grade:
                        # 获取grade列的索引位置并转换为整数
                        grade_col_idx = df.columns.get_loc('grade')
                        if isinstance(grade_col_idx, (int, float)):
                            # 将索引转换为整数后再进行加法运算
                            cell = worksheet.cell(row=int(idx) + 2, column=int(grade_col_idx) + 1)
                        else:
                            # 如果get_loc返回的不是整数，则抛出异常
                            raise ValueError("无法获取grade列的索引位置")
                        # 导入openpyxl模块
                        cell.fill = PatternFill(
                            start_color=GradeCalculator.get_grade_color(grade).lstrip('#'),
                            end_color=GradeCalculator.get_grade_color(grade).lstrip('#'),
                            fill_type='solid'
                        )
            
            return FileResponse(
                temp_file.name,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename='student_score_summary.xlsx',
                background=BackgroundTask(lambda: os.unlink(temp_file.name))
            )
    
    except Exception as e:
        logger.error(f"导出学生成绩汇总失败: {str(e)}")
        raise HTTPException(status_code=500, detail="导出失败")


@router.post("/export-summary")
def export_student_score_summary_to_excel(filter: RecordFilter, db: Session = Depends(get_db)):
    """
    导出学生成绩汇总为Excel
    """
    # 获取汇总数据
    summary = get_student_score_summary(filter, db)
    
    if not summary:
        raise HTTPException(status_code=404, detail="未找到符合条件的记录")
    
    # 创建DataFrame
    df = pd.DataFrame(summary)
    
    # 创建临时文件保存Excel
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        df.to_excel(temp_file.name, index=False)
        temp_file_path = temp_file.name
    
    # 返回文件
    response = FileResponse(
        temp_file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"学生成绩汇总_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    )
    
    # 设置回调函数，在响应发送后删除临时文件
    response.background = BackgroundTask(os.unlink, temp_file_path)
    
    return response


@router.post("/subject-summary")
def get_subject_score_summary_by_date(filter: RecordFilter, db: Session = Depends(get_db)):
    """
    获取指定学科的作业记录汇总，按日期分组
    """
    if not filter.subject:
        raise HTTPException(status_code=400, detail="必须指定学科")
    
    # 构建基础查询
    query = db.query(
        Record.student_id,
        Student.name,
        Student.grade,
        Student.class_name,
        Record.date,
        Record.score
    ).join(Student)
    
    # 应用筛选条件
    query = query.filter(Record.subject == filter.subject)
    if filter.grade:
        query = query.filter(Student.grade == filter.grade)
    if filter.class_name:
        query = query.filter(Student.class_name == filter.class_name)
    if filter.group:
        query = query.filter(Student.group == filter.group)
    if filter.start_date:
        query = query.filter(Record.date >= filter.start_date)
    if filter.end_date:
        query = query.filter(Record.date <= filter.end_date)
    
    # 执行查询
    results = query.all()
    
    # 如果没有记录，返回空列表
    if not results:
        return []
    
    # 获取所有日期
    dates = sorted(set(r.date for r in results))
    
    # 按学生ID分组结果
    student_scores = {}
    for r in results:
        if r.student_id not in student_scores:
            student_scores[r.student_id] = {
                "student_id": r.student_id,
                "name": r.name,
                "grade": r.grade,
                "class_name": r.class_name,
                "total_score": 0,
                "count": 0
            }
        
        # 添加日期分数
        date_str = r.date.strftime("%Y-%m-%d")
        student_scores[r.student_id][date_str] = r.score
        
        # 更新总分和计数
        if r.score is not None:
            student_scores[r.student_id]["total_score"] += r.score
            student_scores[r.student_id]["count"] += 1
    
    # 转换为列表并添加ID
    summary = []
    for i, (_, score) in enumerate(student_scores.items(), 1):
        score["id"] = i
        summary.append(score)
    
    return summary


@router.post("/export-subject-summary")
def export_subject_score_summary_to_excel(filter: RecordFilter, db: Session = Depends(get_db)):
    """
    导出学科作业记录汇总为Excel
    """
    # 获取汇总数据
    summary = get_subject_score_summary_by_date(filter, db)
    
    if not summary:
        raise HTTPException(status_code=404, detail="未找到符合条件的记录")
    
    # 创建DataFrame
    df = pd.DataFrame(summary)
    
    # 创建临时文件保存Excel
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        df.to_excel(temp_file.name, index=False)
        temp_file_path = temp_file.name
    
    # 创建后台任务用于删除临时文件
    background_task = BackgroundTask(os.unlink, temp_file_path)
    
    # 返回文件，并设置后台任务
    return FileResponse(
        temp_file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{filter.subject}学科作业汇总_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
        background=background_task
    )


@router.get("/grades", response_model=List[str])
def get_grades_by_date_range(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """根据日期范围获取年级列表"""
    try:
        stmt = select(Student.grade).distinct()
        stmt = stmt.join(Record)
        
        if start_date:
            stmt = stmt.where(Record.date >= start_date)
        if end_date:
            stmt = stmt.where(Record.date <= end_date)
            
        grades = db.execute(stmt).scalars().all()
        return [grade for grade in grades if grade is not None]
    except SQLAlchemyError as e:
        logger.error(f"获取年级列表失败: {str(e)}")
        raise HTTPException(status_code=500, detail="数据库查询失败")

@router.get("/classes", response_model=List[str])
def get_classes_by_grade_and_date(
    grade: str,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """根据年级和日期范围获取班级列表"""
    try:
        stmt = select(Student.class_name).distinct()
        stmt = stmt.join(Record)
        stmt = stmt.where(Student.grade == grade)
        
        if start_date:
            stmt = stmt.where(Record.date >= start_date)
        if end_date:
            stmt = stmt.where(Record.date <= end_date)
            
        classes = db.execute(stmt).scalars().all()
        return [class_name for class_name in classes if class_name is not None]
    except SQLAlchemyError as e:
        logger.error(f"获取班级列表失败: {str(e)}")
        raise HTTPException(status_code=500, detail="数据库查询失败")